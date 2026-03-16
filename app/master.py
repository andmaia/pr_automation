"""
app/master.py
─────────────────────────────────────────────────────────────────
Baixa a planilha mestre do OneDrive, faz append do DataFrame
e re-envia — tudo em memória (sem tocar o disco do Render).
"""

import os
import httpx
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
from dotenv import load_dotenv

load_dotenv()

# ── variáveis de ambiente (configure no painel do Render) ────────────────────
CLIENT_ID     = os.getenv("MS_CLIENT_ID")
CLIENT_SECRET = os.getenv("MS_CLIENT_SECRET")
TENANT_ID     = os.getenv("MS_TENANT_ID", "consumers")
# Caminho dentro do OneDrive, ex: /me/drive/root:/Financeiro/dados_master.xlsx
ONEDRIVE_PATH = os.getenv("ONEDRIVE_FILE_PATH", "/me/drive/root:/Financeiro/dados_master.xlsx")
ABA           = os.getenv("MASTER_ABA", "Lançamentos")   # aba onde os dados ficam


# ── autenticação ─────────────────────────────────────────────────────────────
def _get_token() -> str:
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    r = httpx.post(url, data={
        "grant_type":    "client_credentials",
        "client_id":     CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope":         "https://graph.microsoft.com/.default",
    }, timeout=30)
    r.raise_for_status()
    return r.json()["access_token"]


# ── download ─────────────────────────────────────────────────────────────────
def _download_xlsx() -> bytes:
    token = _get_token()
    url   = f"https://graph.microsoft.com/v1.0{ONEDRIVE_PATH}:/content"
    r = httpx.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=60)
    r.raise_for_status()
    return r.content


# ── upload ───────────────────────────────────────────────────────────────────
def _upload_xlsx(conteudo: bytes) -> None:
    token = _get_token()
    url   = f"https://graph.microsoft.com/v1.0{ONEDRIVE_PATH}:/content"
    r = httpx.put(
        url,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type":  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
        content=conteudo,
        timeout=60,
    )
    r.raise_for_status()


# ── função principal ─────────────────────────────────────────────────────────
def append_to_master_onedrive(novos_dados: pd.DataFrame) -> dict:
    """
    1. Baixa o xlsx atual do OneDrive para memória.
    2. Encontra a última linha com dados na aba correta.
    3. Valida e reordena colunas do DataFrame para bater com o cabeçalho.
    4. Escreve as novas linhas preservando toda a formatação existente.
    5. Re-envia o arquivo para o OneDrive.

    Retorna {"linhas_inseridas": N}
    """

    # 1. baixar
    conteudo_original = _download_xlsx()
    wb = load_workbook(BytesIO(conteudo_original))

    if ABA not in wb.sheetnames:
        raise ValueError(
            f"Aba '{ABA}' não encontrada. Abas disponíveis: {wb.sheetnames}"
        )

    ws = wb[ABA]

    # 2. localizar primeira linha vazia real (ignora linhas totalmente vazias no fim)
    ultima_linha_com_dado = 1
    for row in range(ws.max_row, 1, -1):
        if any(ws.cell(row=row, column=c).value is not None
               for c in range(1, ws.max_column + 1)):
            ultima_linha_com_dado = row
            break
    primeira_vazia = ultima_linha_com_dado + 1

    # 3. ler cabeçalho e reordenar colunas
    cabecalho = [
        ws.cell(row=1, column=c).value
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=c).value is not None
    ]

    colunas_faltando = set(cabecalho) - set(novos_dados.columns)
    if colunas_faltando:
        raise ValueError(
            f"Colunas ausentes no DataFrame extraído: {colunas_faltando}. "
            f"Colunas disponíveis: {list(novos_dados.columns)}"
        )

    novos_dados = novos_dados[cabecalho]   # garante a ordem certa

    # 4. escrever linha a linha (preserva estilos das células existentes)
    for i, row in enumerate(novos_dados.itertuples(index=False), start=primeira_vazia):
        for j, valor in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=valor)

    # 5. salvar em memória e fazer upload
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    _upload_xlsx(output.read())

    return {"linhas_inseridas": len(novos_dados)}
