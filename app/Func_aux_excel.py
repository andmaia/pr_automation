import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

def create_dataframe(grouped_transactions):
    df = pd.DataFrame({
        'Data': [sublist[0] for sublist in grouped_transactions],
        'descricao': [sublist[2]  for sublist in grouped_transactions ],
        "Valor": [float(sublist[3].replace('.', '', sublist[3].count('.') - 2).replace(',', '.').replace('D', '').replace('C', '')) for sublist in grouped_transactions],
        "Forma pagamento": [ sublist[1]for sublist in grouped_transactions],
        "Tipo": [sublist[4] for sublist in grouped_transactions],
        "CPF": [sublist[-3] for sublist in grouped_transactions],
        "Nome": [sublist[5] for sublist in grouped_transactions],
        "CNPJ":[sublist[-2] for sublist in grouped_transactions],
        "Obs": [sublist[-1] for sublist in grouped_transactions]
    })

    return df

def save_as_excel(df, filename):
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Lançamentos', index=False)

        workbook = writer.book
        sheet = writer.sheets['Lançamentos']

        sheet.column_dimensions['A'].number_format = 'dd/mm/yy'
        sheet.column_dimensions['C'].number_format = '#,##0.00'

        for cell in sheet['1']:
            cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            cell.font = Font(color='FFFFFF')

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            tipo_cell = row[4]  
            valor_cell = row[2]  
            if tipo_cell.value == 'Recebimento':
                valor_cell.font = Font(color='339933')  
                tipo_cell.font = Font(color='339933')  
            elif tipo_cell.value == 'Pagamento':
                valor_cell.font = Font(color='CC3333')  
                tipo_cell.font = Font(color='CC3333')  
        
        table = Table(displayName="Table1", ref=sheet.dimensions)
        style = TableStyleInfo(name="TableStyleLight8", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        sheet.add_table(table)

        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        workbook.save(filename)