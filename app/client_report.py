"""
app/client_report.py
══════════════════════════════════════════════════════════════════════════════
  PLANILHA DO CLIENTE — CONFIGURAÇÃO E GERAÇÃO
══════════════════════════════════════════════════════════════════════════════
"""

import os
import re
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ══════════════════════════════════════════════════════════════════════════════
#  PARTE 1 · CONFIGURAÇÃO  ↓↓↓  EDITE AQUI  ↓↓↓
# ══════════════════════════════════════════════════════════════════════════════

EMPRESA = "Peça Rara"

# Mês de referência — lido da variável de ambiente APP_MES_REF
# Configure no Render: APP_MES_REF=Março/2026
# Pode ser sobrescrito por parâmetro na chamada de gerar_relatorio_cliente()
MES_REF_PADRAO = os.getenv("APP_MES_REF", "Fevereiro/2026")

SALDO_INICIAL_BANCO  = 0.0
SALDO_CAIXA_ANT      = 0.0
RETIRADA_FUNDO       = 0.0
VALORES_MES_ANTERIOR = {"credito": 0.0, "debito": 0.0}
FUNDO_EMERGENCIAL    = {"credito": 0.0}

# ── Sócios — retirada sócio ───────────────────────────────────────────────────
# Lançamentos de PAGAMENTO cujo CPF, CNPJ, Nome ou OBS contenha qualquer
# dos termos abaixo são classificados como "Retirada Sócio" na DRE.
# A busca é case-insensitive e verifica todos os campos do extrato.
# Inclui também lançamentos das Sangrias que batam com esses termos.
#
SOCIOS = [
    "ranielle",
    "raniele",    # variação de grafia com 1 L
    "assem",
    # adicione outros sócios aqui
]

# Palavras-chave adicionais para sangrias de sócio
# Busca no Complemento da sangria — cobre apelidos e variações
PALAVRAS_SANGRIA_SOCIO = [
    "retirada raniel",   # cobre raniele, ranielle, ranieldo...
    "retirada rani",
    "dona rani",
    # adicione outros padrões aqui
]

# ── Palavras-chave de exclusão ────────────────────────────────────────────────
# Lançamentos de SAÍDA cujo OBS ou Descrição contenha a palavra
# são marcados com a categoria indicada e EXCLUÍDOS dos cálculos de Saída/DRE.
#
PALAVRAS_CHAVE = {
    "deposito":              "Depósito em conta",
    "deposição":             "Depósito em conta",
    "transferencia interna": "Transferência Interna",
    "troco":                 "Transferência Interna",
}

CORES = {
    "header_titulo":  "1F3864",
    "header_coluna":  "2F5496",
    "header_texto":   "FFFFFF",
    "secao_fundo":    "D6E4F0",
    "secao_texto":    "1F3864",
    "linha_par":      "EBF3FB",
    "linha_impar":    "FFFFFF",
    "total_fundo":    "BDD7EE",
    "positivo":       "1F7A1F",
    "negativo":       "C00000",
    "input_fundo":    "FFF2CC",
    "input_texto":    "7F6000",
    "link_formula":   "008000",
    "dre_receita":    "1A5276",
    "dre_custo":      "922B21",
    "dre_resultado":  "1E8449",
    "dre_socio":      "6C3483",   # roxo — retirada sócio
    "dados_cat":      "F5EEF8",
    "socio_fundo":    "F5EEF8",   # fundo lilás para linhas de sócio
}

FONTE = "Arial"


# ══════════════════════════════════════════════════════════════════════════════
#  PARTE 2 · GERAÇÃO
# ══════════════════════════════════════════════════════════════════════════════

# ── helpers ───────────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(color="000000", bold=False, size=10, italic=False):
    return Font(name=FONTE, color=color, bold=bold, size=size, italic=italic)

def _border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _border_medium():
    s = Side(style="medium", color="2F5496")
    return Border(left=s, right=s, top=s, bottom=s)

def _alinhar(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _titulo_aba(ws, texto, n_cols=9, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=texto)
    c.fill = _fill(CORES["header_titulo"])
    c.font = _font(CORES["header_texto"], bold=True, size=12)
    c.alignment = _alinhar("center")
    ws.row_dimensions[row].height = 22

def _subtitulo(ws, texto, row, n_cols=2, col=1):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + n_cols - 1)
    c = ws.cell(row=row, column=col, value=texto)
    c.fill = _fill(CORES["secao_fundo"])
    c.font = _font(CORES["secao_texto"], bold=True)
    c.alignment = _alinhar("left")
    ws.row_dimensions[row].height = 16

def _header_cols(ws, headers, row, col_inicio=1, larguras=None):
    for i, h in enumerate(headers, start=col_inicio):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = _fill(CORES["header_coluna"])
        c.font = _font(CORES["header_texto"], bold=True)
        c.alignment = _alinhar("center")
        c.border = _border_thin()
        if larguras:
            ws.column_dimensions[get_column_letter(i)].width = larguras[i - col_inicio]
    ws.row_dimensions[row].height = 16

def _linha_dados(ws, valores, row, col_inicio=1, fundo=None):
    bg = fundo or (CORES["linha_par"] if row % 2 == 0 else CORES["linha_impar"])
    for i, v in enumerate(valores, start=col_inicio):
        c = ws.cell(row=row, column=i, value=v)
        c.fill = _fill(bg); c.font = _font(size=10)
        c.border = _border_thin(); c.alignment = _alinhar("left")

def _total(ws, rotulo, formula_ou_valor, row, col_rot=1, col_val=2, merge=None):
    if merge and merge > 1:
        ws.merge_cells(start_row=row, start_column=col_rot,
                       end_row=row, end_column=col_rot + merge - 1)
    for col, val in [(col_rot, rotulo), (col_val, formula_ou_valor)]:
        c = ws.cell(row=row, column=col, value=val)
        c.fill = _fill(CORES["total_fundo"])
        c.font = _font(bold=True)
        c.alignment = _alinhar("right")
        c.border = _border_thin()
        if col == col_val:
            c.number_format = 'R$ #,##0.00;(R$ #,##0.00);"-"'

def _row_balanco(ws, linha_ref, label, valor_ou_formula,
                 is_input=False, is_total=False, cor=None):
    c_l = ws.cell(row=linha_ref[0], column=1, value=label)
    c_l.font = _font(bold=is_total); c_l.alignment = _alinhar("left")
    c_l.border = _border_thin()
    c_v = ws.cell(row=linha_ref[0], column=2, value=valor_ou_formula)
    c_v.number_format = 'R$ #,##0.00;(R$ #,##0.00);"-"'
    c_v.alignment = _alinhar("right"); c_v.border = _border_thin()
    if is_input:
        for c in (c_l, c_v): c.fill = _fill(CORES["input_fundo"])
        c_v.font = _font(CORES["input_texto"], bold=True)
    elif is_total:
        for c in (c_l, c_v): c.fill = _fill(CORES["total_fundo"])
        c_v.font = _font(bold=True)
    else:
        c_v.font = _font(cor or CORES["link_formula"])
    ret = linha_ref[0]; linha_ref[0] += 1
    return ret


# ── Classificadores ───────────────────────────────────────────────────────────

def _campos_texto(row: pd.Series) -> str:
    """Concatena todos os campos de texto de uma linha para busca."""
    campos = ["descricao", "Obs", "Nome", "CPF", "CNPJ", "Forma pagamento"]
    partes = []
    for c in campos:
        v = row.get(c, "")
        if v and str(v).strip():
            partes.append(str(v).strip())
    return " ".join(partes).lower()


def _eh_socio(row: pd.Series, socios: list) -> bool:
    """Retorna True se o lançamento pertence a um sócio."""
    texto = _campos_texto(row)
    return any(s.lower() in texto for s in socios)


def _classificar_extrato(
    df: pd.DataFrame,
    socios: list,
    palavras_chave: dict,
) -> pd.DataFrame:
    """
    Adiciona coluna '_flag' ao DataFrame:
      - ""         → lançamento normal (entra em tudo)
      - "socio"    → retirada sócio (entra na DRE como Retirada Sócio, não em Saídas)
      - "excluido" → depósito/transferência interna (não entra em nenhum cálculo)

    Só lançamentos de Pagamento são avaliados.
    Sócio tem prioridade sobre palavra-chave.
    """
    df = df.copy()
    df["_flag"]     = ""
    df["_categoria"] = ""

    tipo_col = "Tipo" if "Tipo" in df.columns else None

    for idx, row in df.iterrows():
        tipo = row.get(tipo_col, "") if tipo_col else ""

        # 1. Sócio — avaliado em QUALQUER tipo (Pagamento ou Recebimento)
        if _eh_socio(row, socios):
            df.at[idx, "_flag"] = "socio"
            # Recebimento de sócio = Auxílio, Pagamento = Retirada
            df.at[idx, "_categoria"] = "Auxílio Sócio" if tipo == "Recebimento" else "Retirada Sócio"
            continue

        # 2. Palavras-chave de exclusão — só Pagamentos
        if tipo != "Pagamento":
            continue
        texto = _campos_texto(row)
        for palavra, categoria in palavras_chave.items():
            if palavra.lower() in texto:
                df.at[idx, "_flag"]     = "excluido"
                df.at[idx, "_categoria"] = categoria
                break

    return df


def _classificar_sangrias(
    df: pd.DataFrame,
    socios: list,
    palavras_sangria_socio: list | None = None,
) -> pd.DataFrame:
    """
    Marca sangrias de sócios com _flag='socio'.
    Verifica Categoria e Complemento contra:
      1. socios        — lista de nomes/termos de sócios
      2. palavras_sangria_socio — padrões extras específicos para sangrias
                                   (ex: "retirada raniel", "dona rani")
    """
    if df is None or df.empty:
        return df
    df = df.copy()
    df["_flag"] = ""
    termos = [s.lower() for s in socios]
    if palavras_sangria_socio:
        termos += [p.lower() for p in palavras_sangria_socio]
    for idx, row in df.iterrows():
        texto = (str(row.get("Categoria", "")) + " " +
                 str(row.get("Complemento", ""))).lower()
        if any(t in texto for t in termos):
            df.at[idx, "_flag"] = "socio"
    return df


# ── ABA: Extrato Bancário ─────────────────────────────────────────────────────

def _aba_extrato(wb, df: pd.DataFrame, mes_ref: str):
    ws = wb.create_sheet("Extrato Bancário")
    _titulo_aba(ws, f"{EMPRESA} — Extrato Bancário  |  {mes_ref}", n_cols=10)

    headers  = ["Data", "Descrição", "Valor", "Forma de pagamento",
                "TIPO", "CPF", "NOME", "CNPJ", "OBS", "Flag"]
    larguras = [12, 38, 14, 18, 14, 16, 28, 18, 22, 12]
    _header_cols(ws, headers, row=2, larguras=larguras)

    cols_df = ["Data", "descricao", "Valor", "Forma pagamento",
               "Tipo", "CPF", "Nome", "CNPJ", "Obs"]
    cols_df = [c for c in cols_df if c in df.columns]

    for i, (idx, row_data) in enumerate(df.iterrows(), start=3):
        flag = row_data.get("_flag", "")
        if flag == "socio":
            bg = CORES["socio_fundo"]
        elif flag == "excluido":
            bg = "FDE8D8"
        else:
            bg = CORES["linha_par"] if i % 2 == 0 else CORES["linha_impar"]

        for j, col in enumerate(cols_df, start=1):
            v = row_data.get(col, "")
            c = ws.cell(row=i, column=j, value=v)
            c.fill = _fill(bg); c.border = _border_thin()
            c.alignment = _alinhar("left"); c.font = _font(size=10)
            if col == "Valor":
                c.number_format = 'R$ #,##0.00;(R$ #,##0.00);"-"'
                tipo = row_data.get("Tipo", "")
                c.font = _font(CORES["positivo"] if tipo == "Recebimento" else CORES["negativo"], size=10)
            elif col == "Data":
                c.number_format = "DD/MM/YYYY"; c.alignment = _alinhar("center")

        # Coluna Flag — entra na Table1 para SUMIFS
        c_flag = ws.cell(row=i, column=len(cols_df) + 1, value=flag)
        c_flag.fill = _fill(bg)
        cor_flag = CORES["dre_socio"] if flag == "socio" else ("7B241C" if flag == "excluido" else "808080")
        c_flag.font = _font(cor_flag, size=9, italic=bool(flag))
        c_flag.border = _border_thin()

    n = len(df)
    if n > 0:
        n_cols_total = len(cols_df) + 1
        ref = f"A2:{get_column_letter(n_cols_total)}{n + 2}"
        tbl = Table(displayName="Table1", ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tbl)
    ws.freeze_panes = "A3"


# ── ABA: Dados ────────────────────────────────────────────────────────────────

def _aba_dados(wb, df: pd.DataFrame, mes_ref: str):
    """
    Aba analítica: Data | Descrição | Valor | Forma Pagamento | OBS | Categoria | Flag
    Categoria pré-preenchida para sócios e palavras-chave.
    Pronta para futura categorização por IA (todos os campos relevantes visíveis).
    """
    ws = wb.create_sheet("Dados")
    _titulo_aba(ws, f"{EMPRESA} — Dados para Análise  |  {mes_ref}", n_cols=7)

    headers  = ["Data", "Descrição", "Valor", "Forma Pagamento", "OBS", "Categoria", "Classificação"]
    larguras = [12, 40, 14, 18, 30, 26, 20]
    _header_cols(ws, headers, row=2, larguras=larguras)

    # nota coluna Categoria
    ws.cell(row=2, column=6).value = "Categoria (preencha / IA)"
    ws.cell(row=2, column=6).fill = _fill(CORES["input_fundo"])
    ws.cell(row=2, column=6).font = _font(CORES["input_texto"], bold=True)

    # Filtra: só Pagamentos (saídas) na aba Dados
    df_dados = df[df.get("Tipo", pd.Series(dtype=str)) == "Pagamento"] if "Tipo" in df.columns else df

    for i, (idx, row_data) in enumerate(df_dados.iterrows(), start=3):
        flag     = row_data.get("_flag", "")
        cat_auto = row_data.get("_categoria", "")

        if flag == "socio":
            bg = CORES["socio_fundo"]
        elif flag == "excluido":
            bg = "FDE8D8"
        else:
            bg = CORES["linha_par"] if i % 2 == 0 else CORES["linha_impar"]

        # Data
        c = ws.cell(row=i, column=1, value=row_data.get("Data", ""))
        c.fill = _fill(bg); c.border = _border_thin()
        c.number_format = "DD/MM/YYYY"; c.alignment = _alinhar("center"); c.font = _font(size=10)

        # Descrição
        c = ws.cell(row=i, column=2, value=row_data.get("descricao", ""))
        c.fill = _fill(bg); c.border = _border_thin(); c.font = _font(size=10)

        # Valor
        v = row_data.get("Valor", 0)
        tipo = row_data.get("Tipo", "")
        c = ws.cell(row=i, column=3, value=v)
        c.fill = _fill(bg); c.border = _border_thin()
        c.number_format = 'R$ #,##0.00;(R$ #,##0.00);"-"'; c.alignment = _alinhar("right")
        c.font = _font(CORES["positivo"] if tipo == "Recebimento" else CORES["negativo"], size=10)

        # Forma Pagamento
        c = ws.cell(row=i, column=4, value=row_data.get("Forma pagamento", ""))
        c.fill = _fill(bg); c.border = _border_thin(); c.alignment = _alinhar("center"); c.font = _font(size=10)

        # OBS
        c = ws.cell(row=i, column=5, value=row_data.get("Obs", ""))
        c.fill = _fill(bg); c.border = _border_thin(); c.font = _font(size=10)

        # Categoria — pré-preenchida ou vazia
        c = ws.cell(row=i, column=6, value=cat_auto)
        c.fill = _fill(CORES["socio_fundo"] if flag == "socio"
                       else ("FDE8D8" if flag == "excluido" else CORES["dados_cat"]))
        c.border = _border_medium()
        c.font = _font(CORES["dre_socio"] if flag == "socio"
                       else ("7B241C" if flag == "excluido" else CORES["input_texto"]), size=10)

        # Tipo / aviso
        aviso = cat_auto if flag == "socio" else ("Excluído DRE" if flag == "excluido" else "")
        c = ws.cell(row=i, column=7, value=aviso)
        c.fill = _fill(bg); c.border = _border_thin()
        c.font = _font(CORES["dre_socio"] if flag == "socio"
                       else ("7B241C" if flag == "excluido" else "808080"), italic=bool(flag), size=9)

    ws.freeze_panes = "A3"


# ── ABA: Fechamento de caixa ──────────────────────────────────────────────────

def _aba_caixa(wb, df: pd.DataFrame, mes_ref: str):
    ws = wb.create_sheet("Fechamento de caixa loja")
    _titulo_aba(ws, f"{EMPRESA} — Fechamento de Caixa  |  {mes_ref}", n_cols=3)
    _header_cols(ws, ["Forma de Pagamento", "Entrada de Caixa", "Data"],
                 row=2, larguras=[28, 18, 14])
    col_fp  = df.columns[0]
    col_val = df.columns[1]
    col_dt  = df.columns[2] if len(df.columns) > 2 else None
    for i, (_, r) in enumerate(df.iterrows(), start=3):
        bg = CORES["linha_par"] if i % 2 == 0 else CORES["linha_impar"]
        _linha_dados(ws, [r[col_fp], r[col_val], r[col_dt] if col_dt else ""], i, fundo=bg)
        ws.cell(row=i, column=2).number_format = 'R$ #,##0.00;(R$ #,##0.00);"-"'
        if col_dt: ws.cell(row=i, column=3).number_format = "DD/MM/YYYY"
    ws.freeze_panes = "A3"


# ── ABA: Sangrias ─────────────────────────────────────────────────────────────

def _aba_sangrias(wb, df: pd.DataFrame, mes_ref: str):
    ws = wb.create_sheet("Sangrias")
    _titulo_aba(ws, f"{EMPRESA} — Sangrias  |  {mes_ref}", n_cols=4)
    _header_cols(ws, ["Categoria", "Complemento", "Data", "Valor"],
                 row=2, larguras=[28, 36, 14, 14])
    for i, (_, r) in enumerate(df.iterrows(), start=3):
        flag = r.get("_flag", "")
        bg = CORES["socio_fundo"] if flag == "socio" else (
             CORES["linha_par"] if i % 2 == 0 else CORES["linha_impar"])
        _linha_dados(ws, [r.get("Categoria",""), r.get("Complemento",""),
                          r.get("Data",""), r.get("Valor",0)], i, fundo=bg)
        ws.cell(row=i, column=3).number_format = "DD/MM/YYYY"
        ws.cell(row=i, column=4).number_format = 'R$ #,##0.00;(R$ #,##0.00);"-"'
        # sócio: linha colorida sem coluna extra
    n = len(df)
    _total(ws, "Total Sangrias", f"=SUM(D3:D{n+2})", row=n+3, col_rot=3, col_val=4)
    ws.freeze_panes = "A3"


# ── ABA: Lançamento de despesas ───────────────────────────────────────────────

def _aba_despesas(wb, df: pd.DataFrame, mes_ref: str):
    ws = wb.create_sheet("Lançamento de despesas")
    _titulo_aba(ws, f"{EMPRESA} — Lançamento de Despesas  |  {mes_ref}", n_cols=6)
    _header_cols(ws, ["Data","Descrição","Valor","OBS","Categoria","Forma Pagamento"],
                 row=2, larguras=[12,36,14,24,24,16])
    cols = ["Data","Descrição","Valor","OBS","Categoria","Forma pagamento"]
    cols = [c for c in cols if c in df.columns]
    for i, (_, r) in enumerate(df.iterrows(), start=3):
        bg = CORES["linha_par"] if i % 2 == 0 else CORES["linha_impar"]
        _linha_dados(ws, [r.get(c,"") for c in cols], i, fundo=bg)
        ws.cell(row=i, column=1).number_format = "DD/MM/YYYY"
        ws.cell(row=i, column=3).number_format = 'R$ #,##0.00;(R$ #,##0.00);"-"'
    n = len(df)
    _total(ws, "Total Despesas", f"=SUM(C3:C{n+2})", row=n+3, col_rot=2, col_val=3)
    ws.freeze_panes = "A3"


# ── ABA: Vendedores ───────────────────────────────────────────────────────────

def _aba_vendedores(wb, df: pd.DataFrame, mes_ref: str):
    ws = wb.create_sheet("Vendedores")
    _titulo_aba(ws, f"{EMPRESA} — Desempenho Vendedores  |  {mes_ref}", n_cols=9)
    _header_cols(ws, ["Nome","QTD","Porcentagem","Valor","Data"],
                 row=2, larguras=[18,10,14,16,14])
    col_nome = "Nome " if "Nome " in df.columns else "Nome"
    df = df.copy(); df["_nome"] = df[col_nome].astype(str)
    for i, (_, r) in enumerate(df.iterrows(), start=3):
        bg = CORES["linha_par"] if i % 2 == 0 else CORES["linha_impar"]
        _linha_dados(ws, [r["_nome"], r.get("QTD",""),
                          r.get("Porcentagem",""), r.get("Valor",0), r.get("Data","")], i, fundo=bg)
        ws.cell(row=i, column=3).number_format = "0.0%"
        ws.cell(row=i, column=4).number_format = 'R$ #,##0.00;(R$ #,##0.00);"-"'
        ws.cell(row=i, column=5).number_format = "DD/MM/YYYY"
    resumo = (df.groupby("_nome").agg(valor_total=("Valor","sum"), qtd_total=("QTD","sum"))
               .reset_index().sort_values("valor_total", ascending=False))
    total_valor = resumo["valor_total"].sum()
    _subtitulo(ws, "Ranking Consolidado", row=2, n_cols=3, col=7)
    _header_cols(ws, ["Vendedor","Valor Vendido","% Participação"], row=3, col_inicio=7, larguras=[18,18,16])
    for i, (_, r) in enumerate(resumo.iterrows(), start=4):
        bg = CORES["linha_par"] if i % 2 == 0 else CORES["linha_impar"]
        pct = r["valor_total"] / total_valor if total_valor else 0
        for col, val in [(7, r.iloc[0]), (8, r["valor_total"]), (9, pct)]:
            c = ws.cell(row=i, column=col, value=val)
            c.fill = _fill(bg); c.font = _font(size=10); c.border = _border_thin()
            c.alignment = _alinhar("left" if col == 7 else "right")
        ws.cell(row=i, column=8).number_format = 'R$ #,##0.00;(R$ #,##0.00);"-"'
        ws.cell(row=i, column=9).number_format = "0.0%"
    tot_row = 4 + len(resumo)
    for col, val, fmt in [(7,"TOTAL",None),(8,total_valor,'R$ #,##0.00;(R$ #,##0.00);"-"'),(9,1.0,"0.0%")]:
        c = ws.cell(row=tot_row, column=col, value=val)
        c.fill = _fill(CORES["total_fundo"]); c.font = _font(bold=True)
        c.border = _border_thin(); c.alignment = _alinhar("right")
        if fmt: c.number_format = fmt
    ws.freeze_panes = "A3"


# ── ABA: Faturamento ──────────────────────────────────────────────────────────

def _aba_faturamento(wb, mes_ref: str):
    ws = wb.create_sheet("Faturamento")
    _titulo_aba(ws, f"{EMPRESA} — Faturamento  |  {mes_ref}", n_cols=2)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    l = [3]
    def _row(label, formula, cor=None):
        ws.cell(row=l[0], column=1, value=label).font = _font(size=10)
        ws.cell(row=l[0], column=1).border = _border_thin()
        c = ws.cell(row=l[0], column=2, value=formula)
        c.number_format = 'R$ #,##0.00;(R$ #,##0.00);"-"'
        c.font = _font(cor or CORES["link_formula"], size=10); c.border = _border_thin()
        ret = l[0]; l[0] += 1; return ret
    _subtitulo(ws, "Entrada em conta bancária", 2, n_cols=2); l[0] = 3
    _header_cols(ws, ["Forma de Pagamento","Valor"], row=l[0]); l[0] += 1
    r_cr  = _row("Crédito",       '=SUMIFS(Table1[Valor],Table1[TIPO],"Recebimento",Table1[Forma de pagamento],"CR")')
    r_deb = _row("Débito",        '=SUMIFS(Table1[Valor],Table1[TIPO],"Recebimento",Table1[Forma de pagamento],"DEB")')
    r_dep = _row("Depósito",      '=SUMIFS(Table1[Valor],Table1[TIPO],"Recebimento",Table1[Forma de pagamento],"DEPÓSITO")')
    r_pix = _row("PIX",           '=SUMIFS(Table1[Valor],Table1[TIPO],"Recebimento",Table1[Forma de pagamento],"PIX")')
    r_ted = _row("Transferência", '=SUMIFS(Table1[Valor],Table1[TIPO],"Recebimento",Table1[Forma de pagamento],"TED")')
    _total(ws, "Subtotal Conta Bancária", f"=SUM(B{r_cr}:B{r_ted})", row=l[0])
    l_sub_banco = l[0]; l[0] += 2
    _subtitulo(ws, "Outras formas de entrada (não entram na conta bancária)", l[0], n_cols=2); l[0] += 1
    _header_cols(ws, ["Forma de Pagamento","Valor"], row=l[0]); l[0] += 1
    r_din = _row("Dinheiro",            "=SUMIF('Fechamento de caixa loja'!A:A,\"A Vista - Dinheiro\",'Fechamento de caixa loja'!B:B)")
    r_cf  = _row("Crédito Fornecedor",  "=SUMIF('Fechamento de caixa loja'!A:A,\"Credito Fornecedor\",'Fechamento de caixa loja'!B:B)")
    r_rf  = _row("Retirada Funcionário","=SUMIF('Fechamento de caixa loja'!A:A,\"Retirada Funcionario\",'Fechamento de caixa loja'!B:B)")
    r_cp  = _row("Cartão Presente",     "=SUMIF('Fechamento de caixa loja'!A:A,\"Voucher\",'Fechamento de caixa loja'!B:B)")
    _total(ws, "Subtotal Outras Entradas", f"=SUM(B{r_din}:B{r_cp})", row=l[0])
    l_sub_caixa = l[0]; l[0] += 2
    _total(ws, "TOTAL GERAL FATURAMENTO", f"=B{l_sub_banco}+B{l_sub_caixa}", row=l[0])
    ws.cell(row=l[0], column=2).font = _font(bold=True, size=11)
    return {"sub_banco": l_sub_banco, "sub_caixa": l_sub_caixa,
            "r_cr": r_cr, "r_deb": r_deb, "r_dep": r_dep, "r_pix": r_pix, "r_ted": r_ted,
            "r_din": r_din, "r_cf": r_cf, "r_rf": r_rf, "r_cp": r_cp}


# ── ABA: Saída ────────────────────────────────────────────────────────────────

def _aba_saida(wb, mes_ref: str, df_sangrias_flag: pd.DataFrame = None):
    ws = wb.create_sheet("Saída")
    _titulo_aba(ws, f"{EMPRESA} — Saídas  |  {mes_ref}", n_cols=2)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    l = [3]
    def _row(label, formula, cor=None):
        ws.cell(row=l[0], column=1, value=label).font = _font(size=10)
        ws.cell(row=l[0], column=1).border = _border_thin()
        c = ws.cell(row=l[0], column=2, value=formula)
        c.number_format = 'R$ #,##0.00;(R$ #,##0.00);"-"'
        c.font = _font(cor or CORES["link_formula"], size=10); c.border = _border_thin()
        ret = l[0]; l[0] += 1; return ret
    _subtitulo(ws, "Saídas conta bancária (excluídas retiradas sócio e depósitos)", 2, n_cols=2); l[0] = 3
    _header_cols(ws, ["Forma de Pagamento","Valor"], row=l[0]); l[0] += 1
    # SUMIFS com Flag<>socio e Flag<>excluido
    r_deb   = _row("Débito",    '=SUMIFS(Table1[Valor],Table1[TIPO],"Pagamento",Table1[Forma de pagamento],"DEB",Table1[Flag],"<>socio",Table1[Flag],"<>excluido")')
    r_ted   = _row("TED",       '=SUMIFS(Table1[Valor],Table1[TIPO],"Pagamento",Table1[Forma de pagamento],"TED",Table1[Flag],"<>socio",Table1[Flag],"<>excluido")')
    r_transf= _row("Transferência PIX", '=SUMIFS(Table1[Valor],Table1[TIPO],"Pagamento",Table1[Forma de pagamento],"TRANSFERÊNCIA",Table1[Flag],"<>socio",Table1[Flag],"<>excluido")')
    r_dev   = _row("Devolução", '=SUMIFS(Table1[Valor],Table1[TIPO],"Pagamento",Table1[Forma de pagamento],"Devolução",Table1[Flag],"<>socio",Table1[Flag],"<>excluido")')
    r_pix   = _row("PIX",       '=SUMIFS(Table1[Valor],Table1[TIPO],"Pagamento",Table1[Forma de pagamento],"PIX",Table1[Flag],"<>socio",Table1[Flag],"<>excluido")')
    _total(ws, "Total de transações", f"=SUM(B{r_deb}:B{r_pix})", row=l[0])
    l_trans = l[0]; l[0] += 1
    r_est   = _row("Estornos recebidos (abate)",
                   '=SUMIFS(Table1[Valor],Table1[Forma de pagamento],"Estorno",Table1[TIPO],"Recebimento")')
    _total(ws, "Total saída conta bancária", f"=B{l_trans}-B{r_est}", row=l[0])
    l_saida_conta = l[0]; l[0] += 2

    # Retirada fundo financeiro (pagamentos para sócio)
    _subtitulo(ws, "Retirada Fundo Financeiro (Sócio)", l[0], n_cols=2); l[0] += 1
    r_socio_banco = _row("Retirada Fundo Financeiro",
                         '=SUMIFS(Table1[Valor],Table1[TIPO],"Pagamento",Table1[Flag],"socio")',
                         cor=CORES["dre_socio"])
    l[0] += 1

    # Sangrias normais (exceto sócio e depósito)
    _subtitulo(ws, "Sangrias (exceto sócio e depósitos)", l[0], n_cols=2); l[0] += 1
    # calcula valor de sangrias de sócio para subtrair
    val_sang_socio = 0.0
    if df_sangrias_flag is not None and not df_sangrias_flag.empty and "_flag" in df_sangrias_flag.columns:
        val_sang_socio = df_sangrias_flag.loc[df_sangrias_flag["_flag"] == "socio", "Valor"].sum()
    r_sang  = _row("Sangrias", '=SUM(Sangrias!D:D)-SUMIFS(Sangrias!D:D,Sangrias!B:B,"*Transferência*")')
    # linha de subtração das sangrias de sócio (valor calculado em Python)
    r_sang_socio = _row("  (−) Sangrias sócio", -val_sang_socio if val_sang_socio else 0,
                        cor=CORES["dre_socio"])
    r_sang_liq = l[0]
    _total(ws, "Total Sangrias (sem sócio)", f"=B{r_sang}+B{r_sang_socio}", row=l[0])
    r_sang_net = l[0]; l[0] += 2

    _total(ws, "TOTAL GERAL DE GASTOS",
           f"=B{l_saida_conta}+B{r_socio_banco}+B{r_sang_net}", row=l[0])
    ws.cell(row=l[0], column=2).font = _font(bold=True, size=11)

    return {
        "saida_conta": l_saida_conta,
        "deb": r_deb, "ted": r_ted, "transf": r_transf,
        "dev": r_dev, "pix": r_pix,
        "trans": l_trans, "est": r_est,
        "socio_banco": r_socio_banco,
        "sang": r_sang, "sang_socio": r_sang_socio, "sang_net": r_sang_net,
        "val_sang_socio": val_sang_socio,
    }


# ── ABA: Balanço mensal ───────────────────────────────────────────────────────

def _aba_balanco(wb, saldo_ini, saldo_cx_ant, retirada_fundo, linhas_saida: dict, mes_ref: str):
    ws = wb.create_sheet("Balanço mensal")
    _titulo_aba(ws, f"{EMPRESA} — Balanço Mensal  |  {mes_ref}", n_cols=2)
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 18
    l = [2]
    def _r(label, val, is_input=False, is_total=False, cor=None):
        return _row_balanco(ws, l, label, val, is_input, is_total, cor)
    _subtitulo(ws, "Conta bancária", l[0], n_cols=2); l[0] += 1
    li = _r("Saldo inicial conta bancária do mês anterior", saldo_ini,    is_input=True)
    le = _r("Entrada conta bancária",   "=SUM(Faturamento!B4:B8)")
    _r("Retirada fundo financeiro (informativo)", retirada_fundo,          is_input=True)
    lg = _r("Gastos conta bancária",    f"=Saída!B{linhas_saida['saida_conta']}")
    _r("Saldo em conta estimado",       f"=B{li}+B{le}-B{lg}",            is_total=True)
    l[0] += 1
    _subtitulo(ws, "Caixa em espécie", l[0], n_cols=2); l[0] += 1
    lca = _r("Caixa em espécie mês anterior", saldo_cx_ant,                is_input=True)
    ldi = _r("Entrada dinheiro",
             "=SUMIF('Fechamento de caixa loja'!A:A,\"A Vista - Dinheiro\",'Fechamento de caixa loja'!B:B)")
    ldp = _r("Depósito de dinheiro para conta bancária",
             '=SUMIFS(Table1[Valor],Table1[TIPO],"Recebimento",Table1[Forma de pagamento],"DEPÓSITO")')
    lsg = _r("Sangrias (total)", "=SUM(Sangrias!D:D)")
    _r("Saldo caixa espécie",    f"=B{lca}+B{ldi}-B{ldp}-B{lsg}",         is_total=True)
    l[0] += 1
    _subtitulo(ws, "Outras entradas", l[0], n_cols=2); l[0] += 1
    _r("Crédito do fornecedor",
       "=SUMIF('Fechamento de caixa loja'!A:A,\"Credito Fornecedor\",'Fechamento de caixa loja'!B:B)")
    _r("Retirada funcionário",
       "=SUMIF('Fechamento de caixa loja'!A:A,\"Retirada Funcionario\",'Fechamento de caixa loja'!B:B)")
    l_aux_socio = _r("Auxílio Sócio (recebimentos)",
       '=SUMIFS(Table1[Valor],Table1[TIPO],"Recebimento",Table1[Flag],"socio")',
       cor=CORES["dre_socio"])
    l[0] += 1

    _subtitulo(ws, "Retirada Fundo Financeiro", l[0], n_cols=2); l[0] += 1
    _r("Retirada Fundo Financeiro (Sócio)",
       f"=Saída!B{linhas_saida['socio_banco']}",
       cor=CORES["dre_socio"])


# ── ABA: Comparativo ──────────────────────────────────────────────────────────

def _aba_comparativo(wb, valores_mes_ant, fundo_emergencial, mes_ref: str):
    """
    Comparativo Banco x Caixa — receitas de vendas (sem socio, dinheiro, credito fornecedor).
    Colunas: Forma | Banco | Mes Ant | Total Conta | Caixa | Diferenca | Aportes Socio
    """
    ws = wb.create_sheet("Comparativo")
    n_cols = 7
    _titulo_aba(ws, f"{EMPRESA} — Comparativo Banco vs Caixa  |  {mes_ref}", n_cols=n_cols)
    ws.column_dimensions["A"].width = 24
    for col_l, w in zip("BCDEFG", [20, 20, 18, 20, 18, 20]):
        ws.column_dimensions[col_l].width = w

    _subtitulo(ws,
        "Receitas de vendas: Extrato Sicoob x Fechamento de Caixa  "
        "(Dinheiro e Credito Fornecedor nao transitam pelo banco — ver Faturamento)",
        2, n_cols=n_cols)

    _header_cols(ws, [
        "Forma de Pagamento",
        "Extrato Banco (Sicoob)",
        "Mes Anterior (carregado)",
        "Total em Conta (B+C)",
        "Fechamento de Caixa",
        "Diferenca (D-E)",
        "Aportes Socio ⚠",
    ], row=3, larguras=[24, 20, 20, 18, 20, 18, 20])
    ws.cell(row=3, column=7).fill = _fill(CORES["socio_fundo"])
    ws.cell(row=3, column=7).font = _font(CORES["dre_socio"], bold=True)

    def _banco(fp):
        return (
            f'=SUMIFS(Table1[Valor],Table1[TIPO],"Recebimento",'
            f'Table1[Forma de pagamento],"{fp}",'
            f'Table1[Flag],"<>socio",Table1[Flag],"<>excluido")')

    def _banco_socio(fp):
        return (
            f'=SUMIFS(Table1[Valor],Table1[TIPO],"Recebimento",'
            f'Table1[Forma de pagamento],"{fp}",'
            f'Table1[Flag],"socio")')

    def _caixa(*termos):
        partes = [
            f'SUMIF(\'Fechamento de caixa loja\'!A:A,"{t}",\'Fechamento de caixa loja\'!B:B)'
            for t in termos
        ]
        return "=" + "+".join(partes)

    formas = [
        ("Crédito (todos)", _banco("CR"),  valores_mes_ant.get("credito", 0.0),
         _caixa("Cartão Crédito À Vista", "Cartão Crédito 2x á 6x", "Cartão Crédito 7x á 12x"),
         _banco_socio("CR")),
        ("Débito",          _banco("DEB"), valores_mes_ant.get("debito", 0.0),
         _caixa("Cartão Débito"), _banco_socio("DEB")),
        ("PIX",             _banco("PIX"), 0.0,
         _caixa("Pix"),    _banco_socio("PIX")),
    ]

    LINHA_INICIO = 4
    linhas_data = []

    for i, (forma, banco_f, vant, caixa_f, socio_f) in enumerate(formas, start=LINHA_INICIO):
        bg = CORES["linha_par"] if i % 2 == 0 else CORES["linha_impar"]
        for col, val, fmt, cor_f, bg_f in [
            (1, forma,   None,                                   "000000",              bg),
            (2, banco_f, 'R$ #,##0.00;(R$ #,##0.00);"-"',       CORES["link_formula"], bg),
            (3, vant if vant else None,
                         'R$ #,##0.00;(R$ #,##0.00);"-"',
                         CORES["input_texto"] if vant else "808080",
                         CORES["input_fundo"] if vant else bg),
            (5, caixa_f, 'R$ #,##0.00;(R$ #,##0.00);"-"',       CORES["link_formula"], bg),
            (7, socio_f, 'R$ #,##0.00;(R$ #,##0.00);"-"',       CORES["dre_socio"],    CORES["socio_fundo"]),
        ]:
            c = ws.cell(row=i, column=col, value=val)
            if fmt: c.number_format = fmt
            c.fill = _fill(bg_f if col not in (3,7) else (CORES["input_fundo"] if (col==3 and vant) else bg_f))
            c.font = _font(cor_f, size=10)
            c.border = _border_thin(); c.alignment = _alinhar("right" if col > 1 else "left")
        # D total
        c = ws.cell(row=i, column=4, value=f"=B{i}+C{i}")
        c.number_format = 'R$ #,##0.00;(R$ #,##0.00);"-"'; c.fill = _fill(bg)
        c.font = _font(bold=True, size=10); c.border = _border_thin(); c.alignment = _alinhar("right")
        # F diferenca
        c = ws.cell(row=i, column=6, value=f"=D{i}-E{i}")
        c.number_format = 'R$ #,##0.00;-R$ #,##0.00;"-"'; c.fill = _fill(bg)
        c.font = _font(bold=True, size=10); c.border = _border_thin(); c.alignment = _alinhar("right")
        linhas_data.append(i)

    row_tot = LINHA_INICIO + len(formas)
    c = ws.cell(row=row_tot, column=1, value="TOTAL VENDAS")
    c.fill = _fill(CORES["total_fundo"]); c.font = _font(bold=True)
    c.border = _border_thin(); c.alignment = _alinhar("right")
    for col in range(2, 8):
        letra = get_column_letter(col)
        c = ws.cell(row=row_tot, column=col,
                    value=f"=SUM({letra}{LINHA_INICIO}:{letra}{row_tot-1})")
        c.number_format = ('R$ #,##0.00;-R$ #,##0.00;"-"' if col == 6
                           else 'R$ #,##0.00;(R$ #,##0.00);"-"')
        c.fill = _fill(CORES["socio_fundo"] if col == 7 else CORES["total_fundo"])
        c.font = _font(CORES["dre_socio"] if col == 7 else "000000", bold=True)
        c.border = _border_thin(); c.alignment = _alinhar("right")

    notas = [
        "Colunas B-F: apenas receitas de VENDAS (aportes do socio excluidos).",
        "Coluna G (lilas): aportes do socio por essa forma — informativo, nao e venda.",
        "Diferenca negativa (F): caixa registrou mais que o banco → taxas ou parcelas pendentes.",
        "Dinheiro e Credito Fornecedor nao estao aqui — ver aba Faturamento para o total.",
        "Celulas amarelas (C): ajuste manual para creditos parcelados carregados do mes anterior.",
    ]
    row_nota = row_tot + 2
    for j, nota in enumerate(notas):
        c = ws.cell(row=row_nota+j, column=1, value="📌 " + nota)
        c.font = _font("595959", italic=True, size=9)
        ws.merge_cells(start_row=row_nota+j, start_column=1, end_row=row_nota+j, end_column=n_cols)

    ws.freeze_panes = "A4"


# ── ABA: Gastos ───────────────────────────────────────────────────────────────

def _aba_gastos(wb, mes_ref: str, df_sangrias=None, df_despesas=None):
    ws = wb.create_sheet("Gastos")
    _titulo_aba(ws, f"{EMPRESA} — Gastos por Categoria  |  {mes_ref}", n_cols=2)
    ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 18
    _header_cols(ws, ["Categoria","Valor Gasto"], row=2)
    if df_sangrias is not None and not df_sangrias.empty and "Categoria" in df_sangrias.columns:
        col_cat, col_val, df_fonte = "Categoria", "Valor", df_sangrias
    elif df_despesas is not None and not df_despesas.empty:
        col_cat = "Categoria" if "Categoria" in df_despesas.columns else df_despesas.columns[0]
        col_val = "Valor"     if "Valor"     in df_despesas.columns else df_despesas.columns[2]
        df_fonte = df_despesas
    else:
        return
    resumo = (df_fonte.groupby(col_cat)[col_val].sum().sort_values(ascending=False).reset_index())
    for i, (_, r) in enumerate(resumo.iterrows(), start=3):
        bg = CORES["linha_par"] if i % 2 == 0 else CORES["linha_impar"]
        ws.cell(row=i, column=1, value=r[col_cat]).font = _font(size=10)
        ws.cell(row=i, column=1).fill = _fill(bg); ws.cell(row=i, column=1).border = _border_thin()
        c = ws.cell(row=i, column=2, value=r[col_val])
        c.number_format = 'R$ #,##0.00;(R$ #,##0.00);"-"'
        c.font = _font(CORES["negativo"], size=10); c.fill = _fill(bg); c.border = _border_thin()
    _total(ws, "Total Geral", f"=SUM(B3:B{len(resumo)+2})", row=len(resumo)+3)


# ── ABA: DRE ──────────────────────────────────────────────────────────────────

def _aba_dre(wb, linhas_fat: dict, linhas_saida: dict, mes_ref: str):
    ws = wb.create_sheet("DRE")
    _titulo_aba(ws, f"{EMPRESA} — DRE  |  {mes_ref}", n_cols=3)
    ws.column_dimensions["A"].width = 44; ws.column_dimensions["B"].width = 20; ws.column_dimensions["C"].width = 14
    l = [2]

    def _bloco(texto):
        _subtitulo(ws, texto, l[0], n_cols=3); l[0] += 1

    def _linha(label, formula, cor=None, bold=False, indent=False):
        prefixo = "    " if indent else ""
        c_l = ws.cell(row=l[0], column=1, value=prefixo + label)
        c_l.font = _font(bold=bold, size=10); c_l.border = _border_thin(); c_l.alignment = _alinhar("left")
        c_v = ws.cell(row=l[0], column=2, value=formula)
        c_v.number_format = 'R$ #,##0.00;-R$ #,##0.00;"-"'
        c_v.font = _font(cor or "000000", bold=bold, size=10)
        c_v.border = _border_thin(); c_v.alignment = _alinhar("right")
        ws.cell(row=l[0], column=3).border = _border_thin()
        ret = l[0]; l[0] += 1; return ret

    def _sep():
        for col in range(1, 4):
            c = ws.cell(row=l[0], column=col)
            c.fill = _fill(CORES["total_fundo"]); c.border = _border_thin()
        l[0] += 1

    # ══════════════════════════════════════════════════════════════════════════
    # REGRAS DE NEGÓCIO (Peça Rara):
    #   RECEITA DE VENDAS = Crédito + Débito + PIX + Transferência + Dinheiro
    #   NÃO É VENDA (não entra no resultado):
    #     - Crédito Fornecedor: troca por mercadoria — informativo
    #     - Cartão Presente:    recebido antecipado, não entregue — informativo
    #     - Auxílio Sócio:      aporte do sócio — não é receita da loja
    #     - Retirada Fundo Fin: devolução ao sócio — não é despesa da loja
    #   DESPESA OPERACIONAL = saídas bancárias + sangrias SEM sócio
    # ══════════════════════════════════════════════════════════════════════════

    # ── 1. Receita Operacional de Vendas ──────────────────────────────────────
    _bloco("1. Receita Operacional de Vendas")
    r_cr   = _linha("Crédito (cartões)",    "=Faturamento!B4",  cor=CORES["dre_receita"], indent=True)
    r_deb  = _linha("Débito (cartões)",     "=Faturamento!B5",  cor=CORES["dre_receita"], indent=True)
    r_pix  = _linha("PIX",                  "=Faturamento!B7",  cor=CORES["dre_receita"], indent=True)
    r_ted  = _linha("Transferência",        "=Faturamento!B8",  cor=CORES["dre_receita"], indent=True)
    r_din  = _linha("Dinheiro (caixa)",     "=Faturamento!B13", cor=CORES["dre_receita"], indent=True)
    r_rec_vendas = _linha("TOTAL RECEITA DE VENDAS",
                          f"=B{r_cr}+B{r_deb}+B{r_pix}+B{r_ted}+B{r_din}",
                          cor=CORES["dre_receita"], bold=True)
    _sep()

    # ── 2. Deduções ───────────────────────────────────────────────────────────
    _bloco("2. Deduções da Receita")
    r_dev_d = _linha("Devoluções",                 f"=Saída!B{linhas_saida['dev']}",
                     cor=CORES["dre_custo"], indent=True)
    r_est_d = _linha("Estornos recebidos (abate)", f"=Saída!B{linhas_saida['est']}",
                     cor=CORES["dre_receita"], indent=True)
    r_ded   = _linha("Total Deduções",             f"=B{r_dev_d}-B{r_est_d}",
                     cor=CORES["dre_custo"], bold=True)
    _sep()
    r_rec_liq = _linha("RECEITA LÍQUIDA DE VENDAS",
                       f"=B{r_rec_vendas}-B{r_ded}",
                       cor=CORES["dre_receita"], bold=True)
    _sep()

    # ── 3. Despesas Operacionais ──────────────────────────────────────────────
    _bloco("3. Despesas Operacionais")
    r_s_deb    = _linha("Saídas Débito",          f"=Saída!B{linhas_saida['deb']}",    cor=CORES["dre_custo"], indent=True)
    r_s_ted    = _linha("Saídas TED",             f"=Saída!B{linhas_saida['ted']}",    cor=CORES["dre_custo"], indent=True)
    r_s_transf = _linha("Saídas Transferência PIX",f"=Saída!B{linhas_saida['transf']}",cor=CORES["dre_custo"], indent=True)
    r_s_pix    = _linha("Saídas PIX",             f"=Saída!B{linhas_saida['pix']}",    cor=CORES["dre_custo"], indent=True)
    r_sang_d   = _linha("Sangrias operacionais",  f"=Saída!B{linhas_saida['sang_net']}",cor=CORES["dre_custo"], indent=True)
    r_total_desp = _linha("TOTAL DESPESAS OPERACIONAIS",
                          f"=B{r_s_deb}+B{r_s_ted}+B{r_s_transf}+B{r_s_pix}+B{r_sang_d}",
                          cor=CORES["dre_custo"], bold=True)
    _sep()

    # ── 4. Resultado Operacional ──────────────────────────────────────────────
    _bloco("4. Resultado")
    r_resultado = _linha("RESULTADO OPERACIONAL",
                         f"=B{r_rec_liq}-B{r_total_desp}",
                         cor=CORES["dre_resultado"], bold=True)
    ws.cell(row=r_resultado, column=2).font = _font(CORES["dre_resultado"], bold=True, size=11)
    _sep()

    # ── 5. Movimentações Não Operacionais (informativos) ─────────────────────
    _bloco("5. Movimentações Não Operacionais  ⚠ informativos — não afetam o resultado")
    _linha("", "", indent=False)  # linha em branco visual
    l[0] -= 1  # desfaz linha em branco

    r_aux   = _linha("(+) Auxílio Sócio — aporte recebido",
                     '=SUMIFS(Table1[Valor],Table1[TIPO],"Recebimento",Table1[Flag],"socio")',
                     cor=CORES["dre_socio"], indent=True)
    r_cf    = _linha("(+) Crédito Fornecedor — troca por mercadoria",
                     "=Faturamento!B14", cor="595959", indent=True)
    r_cp    = _linha("(+) Cartão Presente — recebimento antecipado",
                     "=Faturamento!B16", cor="595959", indent=True)
    r_ret_b = _linha("(−) Retirada Fundo Financeiro — conta bancária",
                     f"=Saída!B{linhas_saida['socio_banco']}", cor=CORES["dre_socio"], indent=True)
    r_ret_s = _linha("(−) Retirada Fundo Financeiro — sangrias caixa",
                     linhas_saida["val_sang_socio"] if linhas_saida["val_sang_socio"] else 0,
                     cor=CORES["dre_socio"], indent=True)
    _sep()

    # notas explicativas
    l[0] += 1
    notas = [
        "📌 RESULTADO OPERACIONAL = lucro/prejuízo real das vendas da loja.",
        "📌 Auxílio Sócio: aporte do sócio na conta — não é receita de venda, não entra no resultado.",
        "📌 Crédito Fornecedor: fornecedor retirou mercadoria e 'pagou' assim — não é dinheiro novo.",
        "📌 Cartão Presente: cliente pagou mas ainda não consumiu — receita futura, não atual.",
        "📌 Retirada Fundo Financeiro: devolução ao sócio do que foi aportado — não é despesa da loja.",
    ]
    for j, nota in enumerate(notas):
        c = ws.cell(row=l[0] + j, column=1, value=nota)
        c.font = _font("595959", italic=True, size=9)
        ws.merge_cells(start_row=l[0]+j, start_column=1, end_row=l[0]+j, end_column=3)


# ══════════════════════════════════════════════════════════════════════════════
#  FUNÇÃO PÚBLICA
# ══════════════════════════════════════════════════════════════════════════════

def gerar_relatorio_cliente(
    df_sicoob:            pd.DataFrame | None = None,
    df_caixa:             pd.DataFrame | None = None,
    df_sangrias:          pd.DataFrame | None = None,
    df_despesas:          pd.DataFrame | None = None,
    df_vendedores:        pd.DataFrame | None = None,
    dados_caixa_texto:    dict | None = None,
    saldo_inicial:        float = None,
    saldo_caixa_ant:      float = None,
    retirada_fundo:       float = None,
    valores_31_12:        dict  = None,
    valores_mes_anterior: dict  = None,
    fundo_emergencial:    dict  = None,
    mes_ref:              str   = None,       # ← parâmetro de mês (calendário frontend)
    socios:               list  = None,       # ← sócios vindos do frontend
) -> BytesIO:

    # Resolve configurações
    _mes_ref   = mes_ref   or MES_REF_PADRAO
    _palavras  = PALAVRAS_CHAVE

    # Monta lista de sócios expandindo variações de grafia automaticamente
    # Ex: "ranielle" → também adiciona "raniele", "raniel"
    # Isso cobre erros de digitação comuns no complemento das sangrias
    raw_socios = socios if socios is not None else SOCIOS
    _socios = []
    for nome in raw_socios:
        _socios.append(nome)
        # variação com 1 L (ranielle → raniele)
        if nome.lower().endswith("lle"):
            _socios.append(nome[:-2])   # ranielle → raniel + e? não — ranielle→raniele
        if "ll" in nome.lower():
            _socios.append(nome.lower().replace("ll", "l"))
        # variação sem acento
        _socios.append(nome.lower().rstrip("e"))  # ranielle → ranill (não útil)
    # limpa duplicatas e strings inúteis (< 4 chars)
    _socios = list({s.lower() for s in _socios if len(s) >= 4})
    saldo_ini  = saldo_inicial    if saldo_inicial    is not None else SALDO_INICIAL_BANCO
    saldo_cx   = saldo_caixa_ant  if saldo_caixa_ant  is not None else SALDO_CAIXA_ANT
    ret_fundo  = retirada_fundo   if retirada_fundo   is not None else RETIRADA_FUNDO
    v_ant      = valores_mes_anterior or valores_31_12 or VALORES_MES_ANTERIOR
    fundo_em   = fundo_emergencial or FUNDO_EMERGENCIAL

    # Classifica extrato (sócio / excluido / normal)
    if df_sicoob is not None and not df_sicoob.empty:
        df_sicoob = _classificar_extrato(df_sicoob, _socios, _palavras)

    # Classifica sangrias (sócio)
    if df_sangrias is not None and not df_sangrias.empty:
        df_sangrias = _classificar_sangrias(df_sangrias, _socios, PALAVRAS_SANGRIA_SOCIO)

    wb = Workbook()
    wb.remove(wb.active)

    # ── abas de dados ─────────────────────────────────────────────────────────
    if df_sicoob is not None and not df_sicoob.empty:
        _aba_extrato(wb, df_sicoob, _mes_ref)
        _aba_dados(wb, df_sicoob, _mes_ref)
    if df_caixa      is not None and not df_caixa.empty:
        _aba_caixa(wb, df_caixa, _mes_ref)
    if df_sangrias   is not None and not df_sangrias.empty:
        _aba_sangrias(wb, df_sangrias, _mes_ref)
    if df_despesas   is not None and not df_despesas.empty:
        _aba_despesas(wb, df_despesas, _mes_ref)
    if df_vendedores is not None and not df_vendedores.empty:
        _aba_vendedores(wb, df_vendedores, _mes_ref)

    # ── abas de resumo ────────────────────────────────────────────────────────
    if df_sicoob is not None and not df_sicoob.empty:
        linhas_fat   = _aba_faturamento(wb, _mes_ref)
        linhas_saida = _aba_saida(wb, _mes_ref,
                                  df_sangrias_flag=df_sangrias)
        _aba_balanco(wb, saldo_ini, saldo_cx, ret_fundo, linhas_saida, _mes_ref)
        _aba_comparativo(wb, v_ant, fundo_em, _mes_ref)
        _aba_dre(wb, linhas_fat, linhas_saida, _mes_ref)

    tem_sangrias = df_sangrias is not None and not df_sangrias.empty
    tem_despesas = df_despesas is not None and not df_despesas.empty
    if tem_sangrias or tem_despesas:
        _aba_gastos(wb, _mes_ref,
                    df_sangrias=df_sangrias if tem_sangrias else None,
                    df_despesas=df_despesas if tem_despesas else None)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf