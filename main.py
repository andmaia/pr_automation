from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
import io, os

# ── seus módulos existentes ──────────────────────────────────────────────────
from app.extractor   import Extractor
from app.converter   import Converter
from app.Func_aux     import group_transactions, filter_transactions, process_transactions
from app.Func_aux_pdf import extract_text_pdf, extract_text_between_keywords, verify_pattern_returned_from_pdf
from app.Func_aux_excel import create_dataframe, save_as_excel
from app.ProviderSicoob import ProviderSicoob

# ── novo módulo de append ────────────────────────────────────────────────────
#from app.master import append_to_master_onedrive

# ── planilha do cliente ──────────────────────────────────────────────────────
from app.client_report import gerar_relatorio_cliente

app = FastAPI(title="PR Automation")

# serve o index.html estático
app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/")
def root():
    return FileResponse("static/index.html")


# ────────────────────────────────────────────────────────────────────────────
# ROTA 1 — Relatório Sicoob  (já existia)
# ────────────────────────────────────────────────────────────────────────────
@app.post("/relatorio/sicoob")
async def relatorio_sicoob(
    file: UploadFile = File(...),
    first_item: str  = Form("SALDO DO DIA"),
    last_item:  str  = Form("SALDO BLOQ.ANTERIOR"),
):
    pdf_bytes = await file.read()

    texto   = extract_text_pdf(pdf_bytes)
    texto   = verify_pattern_returned_from_pdf(texto)
    transactions = extract_text_between_keywords(texto, first_item, last_item)

    pattern_inicial = r"^\d{2}/\d{2}$"
    exclusion_words = ["Saldo", "SALDO", "saldo"]

    transactions = filter_transactions(transactions, exclusion_words)
    grouped      = group_transactions(transactions, 3, pattern_inicial)

    provider = ProviderSicoob(
        pattern_inicial=pattern_inicial,
        group_words_clear=exclusion_words,
    )
    processed = process_transactions(grouped, provider)

    df = create_dataframe(processed)

    # salva em memória
    buf = io.BytesIO()
    _save_df_to_buffer(df, buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=relatorio_sicoob.xlsx"},
    )


# ────────────────────────────────────────────────────────────────────────────
# ROTA 2 — Extrair caixa  (já existia)
# ────────────────────────────────────────────────────────────────────────────
@app.post("/extrair")
async def extrair(
    files: list[UploadFile] = File(...),
    include_date:  str = Form("off"),
    dashed_lines:  str = Form("off"),
):
    extractor = Extractor()
    files_bytes = [await f.read() for f in files]

    spp_list, dsc_list, spf_list = extractor.extract_all(
        files_bytes,
        include_date=(include_date == "on"),
    )

    def fmt(lst):
        out = []
        for block in lst:
            if isinstance(block, str):
                for line in block.splitlines():
                    if line.strip():
                        out.append(line)
            else:
                out.append(str(block))
        return out

    return JSONResponse({
        "spp": fmt(spp_list),
        "dsc": fmt(dsc_list),
        "spf": fmt(spf_list),
    })


# ────────────────────────────────────────────────────────────────────────────
# ROTA 3 — Exportar caixa como Excel  (já existia)
# ────────────────────────────────────────────────────────────────────────────
@app.post("/exportar")
async def exportar(
    files: list[UploadFile] = File(...),
    include_date: str = Form("off"),
    dashed_lines: str = Form("off"),
):
    extractor = Extractor()
    converter = Converter()
    files_bytes = [await f.read() for f in files]

    spp_list, dsc_list, spf_list = extractor.extract_all(
        files_bytes,
        include_date=(include_date == "on"),
    )

    buf = io.BytesIO()
    converter.to_excel(
        {
            "Vendas por Pagamento": spp_list,
            "Descontos":            dsc_list,
            "Vendas por Func.":     spf_list,
        },
        filename=buf,          # Converter.to_excel aceita BytesIO
    )
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=baixa_de_caixa.xlsx"},
    )


# ────────────────────────────────────────────────────────────────────────────
# ROTA 4 — NOVA: append na planilha mestre no OneDrive
# ────────────────────────────────────────────────────────────────────────────


# ── helper interno ───────────────────────────────────────────────────────────
def _save_df_to_buffer(df, buf: io.BytesIO):
    """Replica save_as_excel mas salva em BytesIO (sem arquivo em disco)."""
    import openpyxl
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.styles import PatternFill, Font
    import pandas as pd

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Lançamentos", index=False)
        wb    = writer.book
        sheet = writer.sheets["Lançamentos"]

        for cell in sheet["1"]:
            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            cell.font = Font(color="FFFFFF")

        for row in sheet.iter_rows(min_row=2):
            tipo_cell  = row[4]
            valor_cell = row[2]
            if tipo_cell.value == "Recebimento":
                valor_cell.font = Font(color="339933")
                tipo_cell.font  = Font(color="339933")
            elif tipo_cell.value == "Pagamento":
                valor_cell.font = Font(color="CC3333")
                tipo_cell.font  = Font(color="CC3333")

        table = Table(displayName="Table1", ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight8",
            showRowStripes=True, showColumnStripes=True,
        )
        sheet.add_table(table)

        for col in sheet.columns:
            w = max(len(str(c.value or "")) for c in col)
            sheet.column_dimensions[col[0].column_letter].width = w + 2


# ────────────────────────────────────────────────────────────────────────────
# ROTA 5 — Planilha do cliente: aba Sicoob
# ────────────────────────────────────────────────────────────────────────────
@app.post("/cliente/sicoob")
async def cliente_sicoob(
    file: UploadFile = File(...),
    first_item: str  = Form("SALDO DO DIA"),
    last_item:  str  = Form("SALDO BLOQ.ANTERIOR"),
):
    pdf_bytes = await file.read()
    texto        = extract_text_pdf(pdf_bytes)
    texto        = verify_pattern_returned_from_pdf(texto)
    transactions = extract_text_between_keywords(texto, first_item, last_item)
    pattern_inicial = r"^\d{2}/\d{2}$"
    exclusion_words = ["Saldo", "SALDO", "saldo"]
    transactions = filter_transactions(transactions, exclusion_words)
    grouped      = group_transactions(transactions, 3, pattern_inicial)
    provider     = ProviderSicoob(pattern_inicial=pattern_inicial, group_words_clear=exclusion_words)
    processed    = process_transactions(grouped, provider)
    df           = create_dataframe(processed)
    buf = gerar_relatorio_cliente(df_sicoob=df, dados_caixa=None)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=relatorio_cliente_sicoob.xlsx"},
    )


# ────────────────────────────────────────────────────────────────────────────
# ROTA 6 — Planilha do cliente: aba Caixa
# ────────────────────────────────────────────────────────────────────────────
@app.post("/cliente/caixa")
async def cliente_caixa(
    files: list[UploadFile] = File(...),
    include_date: str = Form("off"),
    dashed_lines: str = Form("off"),
):
    extractor   = Extractor()
    files_bytes = [await f.read() for f in files]
    dados = extractor.extract_all_as_dataframes(files_bytes, include_date=(include_date == "on"))
    buf = gerar_relatorio_cliente(
        df_caixa     = dados["caixa"],
        df_sangrias  = dados["sangrias"],
        df_vendedores= dados["vendedores"],
    )
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=relatorio_cliente_caixa.xlsx"},
    )


# ────────────────────────────────────────────────────────────────────────────
# ROTA 7 — Planilha completa do cliente (todas as abas)
# Recebe o PDF do Sicoob + PDFs de caixa em um único request.
# Os inputs manuais (saldo_inicial etc.) vêm como campos do form
# e têm como default os valores definidos na Parte 1 do client_report.py.
# ────────────────────────────────────────────────────────────────────────────
@app.post("/cliente/completo")
async def cliente_completo(
    file_sicoob:     UploadFile       = File(...),
    files_caixa:     list[UploadFile] = File(default=[]),
    first_item:      str   = Form("SALDO DO DIA"),
    last_item:       str   = Form("SALDO BLOQ.ANTERIOR"),
    saldo_inicial:   float = Form(0.0),
    saldo_caixa_ant: float = Form(0.0),
    retirada_fundo:  float = Form(0.0),
    credito_31_12:   float = Form(0.0),
    debito_31_12:    float = Form(0.0),
    mes_ref:         str   = Form(""),        # ex: "Março/2026"
    socios_raw:      str   = Form(""),        # "ranielle,assem" — separado por vírgula
):
    # ── Sicoob ────────────────────────────────────────────────────────────────
    pdf_bytes    = await file_sicoob.read()
    texto        = extract_text_pdf(pdf_bytes)
    texto        = verify_pattern_returned_from_pdf(texto)
    transactions = extract_text_between_keywords(texto, first_item, last_item)
    pattern_inicial = r"^\d{2}/\d{2}$"
    exclusion_words = ["Saldo", "SALDO", "saldo"]
    transactions = filter_transactions(transactions, exclusion_words)
    grouped      = group_transactions(transactions, 3, pattern_inicial)
    provider     = ProviderSicoob(pattern_inicial=pattern_inicial, group_words_clear=exclusion_words)
    df_sicoob    = create_dataframe(process_transactions(grouped, provider))

    # ── Fechamento de caixa, sangrias e vendedores ────────────────────────────
    df_caixa = df_sangrias = df_vendedores = None
    if files_caixa:
        extractor   = Extractor()
        files_bytes = [await f.read() for f in files_caixa]
        dados = extractor.extract_all_as_dataframes(files_bytes, include_date=True)
        df_caixa      = dados["caixa"]      if not dados["caixa"].empty      else None
        df_sangrias   = dados["sangrias"]   if not dados["sangrias"].empty   else None
        df_vendedores = dados["vendedores"] if not dados["vendedores"].empty else None

    _mes_ref = mes_ref.strip() or None
    _socios  = [s.strip() for s in socios_raw.split(",") if s.strip()] or None

    buf = gerar_relatorio_cliente(
        df_sicoob         = df_sicoob,
        df_caixa          = df_caixa,
        df_sangrias       = df_sangrias,
        df_vendedores     = df_vendedores,
        saldo_inicial     = saldo_inicial,
        saldo_caixa_ant   = saldo_caixa_ant,
        retirada_fundo    = retirada_fundo,
        valores_31_12     = {"credito": credito_31_12, "debito": debito_31_12},
        mes_ref           = _mes_ref,
        socios            = _socios,
    )
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=relatorio_cliente_completo.xlsx"},
    )